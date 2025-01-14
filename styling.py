from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def apply_styles(ws):
    # Set header font to bold and center the text
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Add borders to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Align the Description column text to the left
    description_column = ws['F']
    for cell in description_column:
        cell.alignment = Alignment(horizontal="left")
