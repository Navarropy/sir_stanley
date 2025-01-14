import sqlite3
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from openpyxl import Workbook, load_workbook
import time
import traceback
from tqdm import tqdm
import re

# Import the styling function
from styling import apply_styles

# Set up Chrome options to disable logs
chrome_options = Options()
chrome_options.add_argument("--log-level=3")  # Suppress logs
# chrome_options.add_argument("--headless")
# Initialize the Chrome WebDriver with suppressed logs
service = Service(service_args=['--silent'])  # Replace with the path to your chromedriver
driver = webdriver.Chrome(service=service, options=chrome_options)

# Set up the SQLite database
conn = sqlite3.connect('scraped_data.db')
cursor = conn.cursor()
cursor.execute('''
    CREATE TABLE IF NOT EXISTS products (
        brand TEXT,
        model TEXT,
        mfg_part TEXT,
        upc TEXT,
        sku TEXT,
        uom TEXT,
        description TEXT,
        UNIQUE(brand, model, mfg_part, upc, sku)
    )
''')
conn.commit()

# Function to check if an item has been scraped
def item_exists(brand, model, mfg_part, upc, sku):
    cursor.execute('''
        SELECT 1 FROM products WHERE brand=? AND model=? AND mfg_part=? AND upc=? AND sku=? LIMIT 1
    ''', (brand, model, mfg_part, upc, sku))
    return cursor.fetchone() is not None

# Function to add a new item to the database
def add_item_to_db(product_info):
    cursor.execute('''
        INSERT OR IGNORE INTO products (brand, model, mfg_part, upc, sku, uom, description)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (product_info["Brand"], product_info["Model"], product_info["Mfg Part"],
          product_info["UPC"], product_info["SKU"], product_info["UOM"], product_info["Description"]))
    conn.commit()

# Function to extract product information
def extract_product_info(product_element):
    try:
        brand = product_element.find_element(By.XPATH, './/*[@class="mfgname"]').text
        model = product_element.find_element(By.XPATH, './/*[contains(text(),"Model")]//span').text
        try:
            mfg_part = product_element.find_element(By.XPATH, './/*[contains(text(),"Mfg Part")]').text.split(":")[-1].strip()
        except:
            breakpoint()
        upc = product_element.find_element(By.XPATH, './/*[contains(text(),"UPC")]').text.split(":")[-1].strip()
        sku = product_element.find_element(By.XPATH, './/*[contains(text(),"SKU")]').text.split(":")[-1].strip()

        uom_div = product_element.find_element(By.XPATH, './/*[@class="ItemCustomSpecs"]')
        uom = uom_div.find_element(By.XPATH, './/span[@class="ProductFormFieldValue"]').text
        extra_uom_text = uom_div.text.replace(f"UOM: {uom}", "").strip()
        uom = f"{uom} {extra_uom_text}" if extra_uom_text else uom

        description = product_element.find_element(By.XPATH, './/*[@class="ProductName"]').text

        return {
            "Brand": brand,
            "Model": model,
            "Mfg Part": mfg_part,
            "UPC": upc,
            "SKU": sku,
            "UOM": uom,
            "Description": description
        }
    except NoSuchElementException:
        print("Some product information could not be extracted.")
        return {}

# Open the target page
url = 'https://b2b.orscanada.com/Store.aspx?SessionCode=989F2AE8B2F84AC49F48B403A418D869&Toc=116494:116494^0^1437|116494^0^3437&PageID=262732030'
driver.get(url)

# Create or load an Excel workbook and sheet
try:
    wb = load_workbook('products_data.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Data"
    ws.append(["Brand", "Model", "Mfg Part #", "UPC", "SKU", "UOM", "Description"])

try:
    brand_elements = driver.find_elements(By.XPATH, "//*[@class='container']//li//a[not(contains(., '*'))]")
    brand_links = [brand.get_attribute('href') for brand in brand_elements]

    for brand_link in tqdm(brand_links, desc="Processing Brands"):
        driver.get(brand_link)
        time.sleep(2)
        
        try:
            WebDriverWait(driver, 1).until(
                EC.element_to_be_clickable((By.XPATH, '//div[@class="container"]//h2[text()="We\'re sorry. We were not able to find a match for your search."]'))
            )
            continue
        except NoSuchElementException:
            print("")
            pass
        except TimeoutException:
            pass

        # Get the total number of items for the brand
        while True:
            try:
                WebDriverWait(driver, 1).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@style="white-space:nowrap;margin-left:20px;"]//b'))
                )
                break
            except:
                pass
        total_items_text = driver.find_element(By.XPATH, '//*[@style="white-space:nowrap;margin-left:20px;"]//b').text
        total_items = int(re.findall(r'\d+', total_items_text)[0])  # Extract the first number

        items_processed = 0  # Counter to track the number of items processed
        continue_scraping = True  # Flag to control the loop

        with tqdm(total=total_items, desc="Processing Products", leave=False) as pbar:
            while continue_scraping:
                try:
                    while continue_scraping:
                        try:
                            WebDriverWait(driver, 1).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@class="itemPriceListLineGridInner"]'))
                            )
                            break
                        except:
                            pass

                    product_elements = driver.find_elements(By.XPATH, '//*[@class="itemPriceListLineGridInner"]')

                    for product_element in product_elements:
                        product_info = extract_product_info(product_element)

                        # Skip items that have already been scraped
                        if product_info and not item_exists(product_info["Brand"], product_info["Model"],
                                                            product_info["Mfg Part"], product_info["UPC"],
                                                            product_info["SKU"]):
                            # Add item to the database
                            add_item_to_db(product_info)

                            # Add item to the Excel sheet
                            ws.append([
                                product_info["Brand"],
                                product_info["Model"],
                                product_info["Mfg Part"],
                                product_info["UPC"],
                                product_info["SKU"],
                                product_info["UOM"],
                                product_info["Description"]
                            ])
                        items_processed += 1
                        pbar.update(1)  # Update the progress bar for each product processed

                    apply_styles(ws)
                    wb.save('products_data.xlsx')

                    next_page_button = driver.find_element(By.XPATH, '//*[@class="ResultsPageLink ResultsPageLinkNext"]')
                    if next_page_button:
                        next_page_button.click()
                        time.sleep(2)
                    else:
                        continue_scraping = False
                except (NoSuchElementException, TimeoutException):
                    continue_scraping = False

except KeyboardInterrupt:
    print("Process interrupted by user.")

except Exception as e:
    print("An error occurred during execution:")
    print(traceback.format_exc())

finally:
    driver.quit()
    wb.save('products_data.xlsx')
    conn.close()
    print("Data saved to products_data.xlsx and database.")
